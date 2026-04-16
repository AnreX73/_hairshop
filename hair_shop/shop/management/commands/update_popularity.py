import os
from django.core.management.base import BaseCommand
from django.conf import settings
from django.db import transaction
from openpyxl import load_workbook
from shop.models import Product  # 👈 замени на своё приложение

class Command(BaseCommand):
    help = 'Обновляет popularity из Excel, остальные товары сбрасывает в 0'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, required=True, help='Имя файла без .xlsx')
        parser.add_argument('--preview', action='store_true', help='Только показать статистику, не сохранять')

    def handle(self, *args, **kwargs):
        filename = kwargs['file']
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        export_dir = os.path.join(settings.BASE_DIR, 'wb_exports')
        file_path = os.path.join(export_dir, filename)

        if not os.path.exists(file_path):
            self.stderr.write(f'❌ Файл не найден: {file_path}')
            return

        self.stdout.write(f'📖 Читаем: {filename}')
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [str(h).strip().lower() for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        try:
            col_art = headers.index('article')
            col_pop = headers.index('popularity')
        except ValueError:
            self.stderr.write('❌ Не найдены колонки "article" или "popularity"')
            wb.close()
            return

        # Собираем валидные данные из Excel
        excel_data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                art = str(row[col_art]).strip()
                pop = int(float(row[col_pop]))
                if art:
                    excel_data[art] = pop
            except (ValueError, TypeError, IndexError):
                pass
        wb.close()

        if not excel_data:
            self.stderr.write('❌ В файле нет валидных данных')
            return

        excel_articles = set(excel_data.keys())

        # 1. Находим товары, которые есть и в Excel, и в БД
        matches = list(Product.objects.filter(article__in=excel_articles))
        for p in matches:
            p.popularity = excel_data[str(p.article).strip()]

        # 2. Кверисет для сброса остальных товаров
        reset_queryset = Product.objects.exclude(article__in=excel_articles)

        if kwargs['preview']:
            self.stdout.write(f'\n👀 ПРЕДПРОСМОТР:')
            self.stdout.write(f'  📦 Уникальных артикулов в Excel: {len(excel_data)}')
            self.stdout.write(f'  ✅ Найдено совпадений в БД (будет обновлено): {len(matches)}')
            self.stdout.write(f'  🔄 Остальные товары (будет сброшено в 0): {reset_queryset.count()}')
            return

        # Сохраняем в одной транзакции
        with transaction.atomic():
            if matches:
                Product.objects.bulk_update(matches, ['popularity'], batch_size=1000)
            
            # Массовый сброс для всех товаров, которых нет в Excel
            reset_count = reset_queryset.update(popularity=0)

            self.stdout.write(self.style.SUCCESS(
                f'✅ Обновлено по файлу: {len(matches)} | Сброшено в 0: {reset_count}'
            ))