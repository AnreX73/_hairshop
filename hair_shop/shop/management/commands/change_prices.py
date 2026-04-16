import os
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from shop.models import Product  # 👈 замени на своё приложение

class Command(BaseCommand):
    help = 'Обновляет цены и обнуляет скидку из указанного файла в wb_exports/'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, required=True, help='Имя файла без расширения .xlsx')

    def handle(self, *args, **kwargs):
        filename = kwargs['file']
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        export_dir = os.path.join(settings.BASE_DIR, 'wb_exports')
        file_path = os.path.join(export_dir, filename)

        if not os.path.exists(file_path):
            self.stderr.write(f'❌ Файл не найден: {file_path}')
            self.stdout.write(f'💡 Проверьте, что он лежит в: {export_dir}')
            return

        self.stdout.write(f'📖 Читаем: {filename}')
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Ищем индексы колонок
        headers = [str(h).strip().lower() for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        try:
            col_art = headers.index('article')
            col_price = headers.index('price')
        except ValueError:
            self.stderr.write('❌ В файле отсутствуют столбцы "article" или "price"')
            return

        # Собираем валидные пары article -> price
        excel_data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                art = str(row[col_art]).strip()
                pr = int(float(row[col_price]))
                if art:
                    excel_data[art] = pr
            except (ValueError, TypeError, IndexError):
                pass

        if not excel_data:
            self.stderr.write('❌ В файле нет валидных данных для обновления')
            return

        # Берём из БД только те товары, чьи артикулы есть в Excel
        products = list(Product.objects.filter(article__in=excel_data.keys()))

        # Применяем изменения
        for p in products:
            p.price = excel_data[str(p.article).strip()]
            p.discount_percentage = 0  # сброс скидки для обновлённых товаров

        if products:
            Product.objects.bulk_update(products, ['price', 'discount_percentage'], batch_size=1000)
            self.stdout.write(self.style.SUCCESS(f'✅ Обновлено {len(products)} товаров'))
        else:
            self.stdout.write('⚠️ Совпадений артикулов с БД не найдено')